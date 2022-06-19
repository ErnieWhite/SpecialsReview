import csv
import openpyxl
import pprint


def get_basis_name(basis_number):
    match_ups = {
        '1': 'LIST',
        '2': 'INTERNAL',
        '3': 'UMSP',
        '4': 'CMP',
        '5': 'STD-COST',
        '6': 'REP-COST',
        '8': 'AVG-COST',
        '9': 'LASTCOST',
        '10': 'CMP',
        '21': 'Lnd Cost',
        '22': 'Avg Lnd',
        '25': 'Ord COGS',
        '28': 'Ord Comm',
        '29': 'Strgc List',
        '30': 'Strgc Cost',
        '31': 'AVG-COST',
    }
    return match_ups.get(basis_number, basis_number)

def setup_workbook(wb):
    wb.create_sheet('CS-Price')
    wb.create_sheet('CS-Cost-Parts')
    wb.create_sheet('CS-Cost-Groups')
    wb.create_sheet('BS-Price')
    wb.create_sheet('BS-Cost-Parts')
    wb.create_sheet('BS-Cost-Groups')
    wb.create_sheet('RC-Price-Parts')
    wb.create_sheet('RC-Price-Parts')
    wb.create_sheet('RC-Cost-Parts')
    wb.create_sheet('RC-Cost-Groups')
    wb.create_sheet('Hierarchy')
    wb.create_sheet('CustomerList')


def update_worksheet(ws, special):
    vals = list(special.values())
    ws.append(vals)


def is_cs_price(special):
    if special['Customer ID'] == '':
        return False
    if special['Home Branch'] == '1000':
        return False
    if special['Price Basis#'] == '':
        return False
    return True


def is_cs_cost_part(special):
    if special['Customer ID'] == '':
        return False
    if special['Home Branch'] == '1000':
        return False
    if special['Cost Basis#'] == '':
        return False
    if not special['PROD_NBR'].isdecimal():
        return False
    return True


def is_cs_cost_group(special):
    if special['Customer ID'] == '':
        return False
    if special['Home Branch'] == '1000':
        return False
    if special['Cost Basis#'] == '':
        return False
    if special['PROD_NBR'].isdecimal():
        return False
    return True


def is_bs_price(special):
    if special['Customer ID'] == '':
        return False
    if special['Home Branch'] != '1000':
        return False
    if special['Price Basis#'] == '':
        return False
    return True


def is_bs_cost_part(special):
    if special['Customer ID'] == '':
        return False
    if special['Home Branch'] != '1000':
        return False
    if special['Cost Basis#'] == '':
        return False
    if not special['PROD_NBR'].isdecimal():
        return False
    return True


def is_bs_cost_group(special):
    if special['Customer ID'] == '':
        return False
    if special['Home Branch'] != '1000':
        return False
    if special['Cost Basis#'] == '':
        return False
    if special['PROD_NBR'].isdecimal():
        return False
    return True


def is_rc_price_part(special):
    if special['Customer ID'] != '':
        return False
    if special['Price Basis#'] == '':
        return False
    if not special['PROD_NBR'].isdecimal():
        return False
    return True


def is_rc_price_group(special):
    if special['Customer ID'] != '':
        return False
    if special['Price Basis#'] == '':
        return False
    if special['PROD_NBR'].isdecimal():
        return False
    return True


def is_rc_cost_part(special):
    if special['Customer ID'] != '':
        return False
    if special['Cost Basis#'] == '':
        return False
    if not special['PROD_NBR'].isdecimal():
        return False
    return True


def is_rc_cost_group(special):
    if special['Customer ID'] != '':
        return False
    if special['Cost Basis#'] == '':
        return False
    if special['PROD_NBR'].isdecimal():
        return False
    return True


def read_special_data(workbooks, f):
    with open(f, 'r', encoding='Windows-1252', newline='') as csv_file:
        specials = csv.DictReader(csv_file)
        for special in specials:
            special['PROD_NBR'] = cleanup_prod_nbr(special['PROD_NBR'])
            branch = special['Home Branch'] if special['Home Branch'] != '1000' else special['Branch/Terr.']
            if branch not in workbooks:
                workbooks[branch] = openpyxl.Workbook()
                setup_workbook(workbooks[branch])
            wb = workbooks[branch]
            if is_cs_price(special):
                update_worksheet(wb['CS-Price'], special)
            if is_cs_cost_part(special):
                update_worksheet(wb['CS-Cost-Parts'], special)
            if is_cs_cost_group(special):
                update_worksheet(wb['CS-Cost-Groups'], special)
            if is_bs_price(special):
                update_worksheet(wb['BS-Price'], special)
            if is_bs_cost_part(special):
                update_worksheet(wb['BS-Cost-Parts'], special)
            if is_bs_cost_group(special):
                update_worksheet(wb['BS-Cost-Groups'], special)


"""
branch
    customer specials
        price   
            parts
            groups
        cost
            parts
            groups
    branch specials
        price
            parts
            groups
        cost
            parts
            groups
    rate cards
        price
            parts
            groups
        cost
            parts
            groups
    hierarchy
    customer list
"""


def cleanup_prod_nbr(prod_nbr: str) -> str:
    """Remove the G from in front of prod_nbr if prod_nbr is a group"""
    if prod_nbr.startswith("G"):
        return prod_nbr[1:]
    return prod_nbr


# def update_contract_specials(specials, special):
#     special['PROD_NBR'] = cleanup_prod_nbr(special['PROD_NBR'])
#     key = ':'.join([special['Branch/Terr.'], special['Customer ID'], special['PROD_NBR']])
#     if 'ContractSpecials' not in specials:
#         specials['ContractSpecials'] = {}
#     if special['Price Basis#']:
#         update_price_specials
#
#
# def update_specials_dict(specials, special):
#     branch = special['Home Branch'] if special['Home Branch'] != '1000' else special['Branch/Terr.']
#     if branch not in specials:
#         specials[branch] = {}
#     if special['Home Branch'] == '1000':
#         return update_contract_specials(specials[branch], special)
#     else:
#         return update_customer_specials(specials[branch], special)
#
#
# def create_specials_dict(specials: list[dict]):
#     specials_dict = dict()
#     while len(specials) > 0:
#         special = specials.pop(0)
#         update_specials_dict(specials, special)
#
#         if special['PROD_NBR'].isdecimal():
#             prod_nbr_type = 'part'
#         else:
#             prod_nbr_type = 'group'
#         if prod_nbr_type not in specials_dict[branch][special_type]:
#             specials_dict[branch][special_type][prod_nbr_type] = {}
#         if key not in specials_dict[branch][special_type][prod_nbr_type]:
#             specials_dict[branch][special_type][prod_nbr_type][key] = {
#                 'Branch/Terr.': special['Branch/Terr.'] if special['Branch/Terr.'] != '' else 'ALL',
#                 'Diff. Matrix Info.': special['Diff. Matrix Info.'],
#                 'Home Branch': special['Home Branch'],
#                 'Job': 'Y' if special['Bill To ID'] == special['Customer ID'] else 'N',
#                 'Bill To ID': special['Bill To ID'] if special['Bill To ID'] != '' else special['Customer ID'],
#                 'Customer ID': special['Customer ID'],
#                 'Customer Name': special['Customer Name'],
#                 'Rate Card': special['Price Class'],
#                 'Outside Salesperson': special['Outside Salesperson'],
#                 'Price Line': special['Price Line'],
#                 'PROD_NBR': special['PROD_NBR'],
#                 'Product Description': special['Product Description'],
#                 'Effective Date': special['Effective Date'],
#                 'Expire Date': special['Expire Date'],
#             }
#             if special['Price Basis#']:
#                 if ('Price Data' not in specials_dict[branch][special_type][prod_nbr_type][key]) or (
#                         special['Matrix ID'] > specials_dict[branch][special_type][prod_nbr_type][key]['Price Data']['Matrix ID']):
#                     specials_dict[branch][special_type][prod_nbr_type][key]['Price Data'] = {
#                         'Matrix ID': special['Matrix ID'],
#                         'SellGroupAll': special['SellGroupAll'],
#                         'SellGroupMscAll': special['SellGroupMscAll'],
#                         'SellGroupMregAll': special['SellGroupMregAll'],
#                         'Price Date Ovrd.': special['Price Date Ovrd.'],
#                         'LIST': float(special['LIST']),
#                         'UMSP': float(special['UMSP']),
#                         'CMP': float(special['CMP']),
#                         'STD-COST': float(special['STD-COST']),
#                         'REP-COST': float(special['REP-COST']),
#                         'AVG-COST': float(special['AVG-COST']),
#                         'LASTCOST': float(special['LASTCOST']),
#                         'Lnd Cost': float(special['Lnd Cost']),
#                         'Avg Lnd': float(special['Avg Lnd']),
#                         'Price Basis': get_basis_name(special['Price Basis#']),
#                         'Price Formula': special['Price Formula'],
#                         'Enable Rnding Rules': special['Enable Rnding Rules'],
#                     }
#
#             if special['Cost Basis#']:
#                 if ('Cost Data' not in specials_dict[branch][special_type][prod_nbr_type][key]) or (
#                         special['Matrix ID'] > specials_dict[branch][special_type][prod_nbr_type][key]['Cost Data']['Matrix ID']):
#                     specials_dict[branch][special_type][prod_nbr_type][key]['Cost Data'] = {
#                         'Matrix ID': special['Matrix ID'],
#                         'Cost Basis': get_basis_name(special['Cost Basis#']),
#                         'Cost Formula': special['Cost Formula'],
#                     }
#     return specials_dict


# def read_rate_card_data(excel_workbooks, param):
#     with open(f, 'r', encoding='Windows-1252', newline='') as csv_file:
#         specials = csv.DictReader(csv_file)
#         for special in specials:
#             special['PROD_NBR'] = cleanup_prod_nbr(special['PROD_NBR'])
#             if branch not in workbooks:
#                 workbooks[branch] = openpyxl.Workbook()
#                 setup_workbook(workbooks[branch])
#             wb = workbooks[branch]
#             if is_cs_price(special):
#                 update_worksheet(wb['CS-Price'], special)
#             if is_cs_cost_part(special):
#                 update_worksheet(wb['CS-Cost-Parts'], special)
#             if is_cs_cost_group(special):
#                 update_worksheet(wb['CS-Cost-Groups'], special)
#             if is_bs_price(special):
#                 update_worksheet(wb['BS-Price'], special)
#             if is_bs_cost_part(special):
#                 update_worksheet(wb['BS-Cost-Parts'], special)
#             if is_bs_cost_group(special):
#                 update_worksheet(wb['BS-Cost-Groups'], special)


def main():
    branch_workbooks = {}
    read_special_data(branch_workbooks, 'data/cs.csv')
    print('cs out')
    read_special_data(branch_workbooks, 'data/bs.csv')
    print('bs out')
    # read_rate_card_data(branch_workbooks, 'data/rc.csv')
    # print('rc out')


    for branch in branch_workbooks:
        branch_workbooks[branch].save(filename=branch + ".xlsx")


    # bs = read_data('data/bs.csv')
    # rc = read_data('data/rc.csv')
    # print(f'CS {len(cs)}')
    # print(f'BS {len(bs)}')
    # print(f'RC {len(rc)}')
    #
    # cs_dict = create_specials_dict(cs)
    # bs_dict = create_specials_dict(bs)
    #
    # print(f'CS Price {len(cs_dict)}')
    # print(f'BS Price {len(bs_dict)}')
    # with open('cs_output.txt', 'wt') as out:
    #     pprint.pprint(cs_dict, stream=out)
    # with open('bs_output.txt', 'wt') as out:
    #     pprint.pprint(bs_dict, stream=out)


if __name__ == "__main__":
    main()
